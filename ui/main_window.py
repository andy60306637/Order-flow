"""
主視窗：整合所有元件與資料流。

布局（深色 QSplitter）：
  ┌─── Toolbar ─────────────────────────────────────────────────┐
  │ [Symbol ▼] [Interval ▼] | [View: Kline|Footprint] [Mode ▼] │
  ├──────────────┬──────────────────────────────────────────────┤
  │              │  QTabWidget: [Kline] [Footprint]             │
  │  Order Book  │  ─────────────────────────────────────────   │
  │  (Level 2)   │  CVD Chart                                   │
  ├──────────────┤                                              │
  │  OB Heatmap  │                                              │
  └──────────────┴──────────────────────────────────────────────┘
"""
from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Optional, List

from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QSplitter, QTabWidget,
    QVBoxLayout, QHBoxLayout, QComboBox, QLabel,
    QFrame, QPushButton, QDockWidget,
    QDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QMessageBox, QDateTimeEdit,
)
from PyQt6.QtCore import QDateTime

from PyQt6.QtGui import QColor
from PyQt6 import QtGui

import config
from core.data_types import Trade, Kline
from core.order_book import OrderBook
from core.cvd_calculator import CvdCalculator
from core.footprint_builder import FootprintBuilder
from core.ws_client import WsWorkerThread
from core.history_processor import HistoryProcessorThread
from core import data_paths, kline_cache, tick_cache
from utils.ui_settings import ui_settings
from strategies import STRATEGY_REGISTRY
from strategies.base import StrategyBase, StrategySignal
from ui.order_book_widget import OrderBookWidget
from ui.kline_chart import KlineChart
from ui.cvd_chart import CvdChart, StatsPanel
from ui.heatmap_widget import HeatmapWidget
from ui.footprint_widget import FootprintChart
from ui.capacity_tab import CapacityTab

logger = logging.getLogger(__name__)

_INTERVAL_MS_MAP = {
    "1m": 60_000, "3m": 180_000, "5m": 300_000, "15m": 900_000,
    "30m": 1_800_000, "1h": 3_600_000, "2h": 7_200_000, "4h": 14_400_000,
    "6h": 21_600_000, "8h": 28_800_000, "12h": 43_200_000, "1d": 86_400_000,
    "3d": 259_200_000, "1w": 604_800_000,
}


def _interval_ms(interval: str) -> int:
    return _INTERVAL_MS_MAP.get(interval, 60_000)


def _format_utc_day(ms: int) -> str:
    return QDateTime.fromMSecsSinceEpoch(int(ms), Qt.TimeSpec.UTC).toString("yyyy-MM-dd")


def _find_snapshot_bar_index(klines: list, ts_ms: int) -> Optional[int]:
    if not klines or not ts_ms:
        return None
    lo, hi = 0, len(klines) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        k = klines[mid]
        if k.open_time <= ts_ms <= k.close_time:
            return mid
        if ts_ms < k.open_time:
            hi = mid - 1
        else:
            lo = mid + 1
    return max(0, min(lo, len(klines) - 1))


def _fallback_snapshot_contexts(trade_list: list[dict], klines: list) -> list[dict]:
    contexts: list[dict] = []
    active_trades = [t for t in trade_list if not t.get("skipped")]
    for idx, trade in enumerate(active_trades):
        entry_ki = _find_snapshot_bar_index(klines, int(trade.get("entry_time", 0) or 0))
        exit_ki = _find_snapshot_bar_index(klines, int(trade.get("exit_time", 0) or 0))
        if entry_ki is None:
            continue
        latest = exit_ki if exit_ki is not None else entry_ki
        contexts.append({
            "trade": trade,
            "trade_idx": idx,
            "k0_signal": None,
            "entry_signal": None,
            "exit_signal": None,
            "k0_ki": None,
            "entry_ki": entry_ki,
            "exit_ki": exit_ki,
            "win_start": max(0, entry_ki - 10),
            "win_end": min(len(klines) - 1, latest + 10),
        })
    return contexts


# ═══════════════════════════════════════════════════════════════════
# 回測結果對話框
# ═══════════════════════════════════════════════════════════════════

class BacktestResultDialog(QDialog):
    """顯示策略回測統計摘要與逐筆交易明細（含市場時區/月份篩選）。"""

    # 各盤口對應的 session key（DST-aware，詳見 _in_session）
    SESSIONS = {
        "全時間": None,
        "亞洲盤": "asia",
        "倫敦盤": "london",
        "紐約盤": "newyork",
    }

    @staticmethod
    def _fmt_pf(v: float) -> str:
        return "∞" if v == float("inf") else f"{v:.2f}"

    @staticmethod
    def _trade_month(t: dict) -> str:
        from datetime import datetime
        ts = t.get("entry_time", 0)
        if not ts:
            return ""
        dt = datetime.fromtimestamp(ts / 1000, tz=config.DISPLAY_TZ)
        return f"{dt.year}-{dt.month:02d}"

    @staticmethod
    def _in_session(t: dict, session_key: str) -> bool:
        """
        判斷交易是否落在指定盤口時段。

        以各市場本地時間判斷（自動處理夏/冬令），換算關係：
          亞洲盤  Asia/Taipei  07:00~16:00（UTC+8，無 DST）
                   → 夏/冬令相同：UTC 23:00~08:00
          倫敦盤  Europe/London 08:00~17:00
                   → 夏令(BST, UTC+1)：UTC 07:00~16:00 ≈ 台灣 15:00~00:00
                   → 冬令(GMT, UTC+0)：UTC 08:00~17:00 ≈ 台灣 16:00~01:00
          紐約盤  America/New_York 08:00~16:00
                   → 夏令(EDT, UTC-4)：UTC 12:00~20:00 ≈ 台灣 20:00~04:00
                   → 冬令(EST, UTC-5)：UTC 13:00~21:00 ≈ 台灣 21:00~05:00
        """
        from datetime import datetime, timezone
        ts = t.get("entry_time", 0)
        if not ts:
            return True
        dt_utc = datetime.fromtimestamp(ts / 1000, tz=timezone.utc)

        try:
            from zoneinfo import ZoneInfo
            if session_key == "asia":
                h = dt_utc.astimezone(ZoneInfo("Asia/Taipei")).hour
                return 7 <= h < 16
            elif session_key == "london":
                h = dt_utc.astimezone(ZoneInfo("Europe/London")).hour
                return 8 <= h < 17
            elif session_key == "newyork":
                h = dt_utc.astimezone(ZoneInfo("America/New_York")).hour
                return 8 <= h < 16
        except Exception:
            # tzdata 未安裝時退回靜態 UTC 範圍（近似）
            _FALLBACK = {
                "asia":    (23, 8),   # UTC 23:00~08:00
                "london":  (8, 17),   # UTC 08:00~17:00（冬令）
                "newyork": (12, 21),  # UTC 12:00~21:00（夏令近似）
            }
            h_range = _FALLBACK.get(session_key)
            if h_range is None:
                return True
            h = dt_utc.hour
            h_start, h_end = h_range
            return h_start <= h < h_end if h_start < h_end else (h >= h_start or h < h_end)

        return True

    # ─────────────────────────────────────────────────────────────────
    def __init__(
        self,
        stats: dict,
        klines: list | None = None,
        tick_map: dict | None = None,
        signals: list | None = None,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._stats = stats
        self._full_trade_list = stats.get("trade_list", [])
        self._klines   = klines   or []
        self._tick_map = tick_map
        self._signals  = signals  or []
        self.setWindowTitle("回測結果")
        self.setMinimumSize(1020, 620)
        self.setStyleSheet(
            f"QDialog {{ background: {config.COLOR_BG}; color: {config.COLOR_FG}; }}"
            f"QTableWidget {{ background: #1e222d; color: {config.COLOR_FG};"
            f" gridline-color: #2a2e39; font-size: 12px; }}"
            f"QHeaderView::section {{ background: #1e222d; color: {config.COLOR_FG};"
            f" border: 1px solid #2a2e39; padding: 4px; font-weight: bold; }}"
        )

        layout = QVBoxLayout(self)

        # ── 回測參數回顯 ─────────────────────────────────────────────
        cap  = stats.get("initial_capital", 0)
        lev  = stats.get("leverage", 0)
        fm   = stats.get("fee_mode", "")
        fr   = stats.get("fee_rate", 0)
        mlp  = stats.get("max_loss_pct", 0)
        slip = stats.get("slippage_bps", 0.0)
        fund = stats.get("funding_rate", 0.0)
        feq  = stats.get("final_equity", 0.0)
        ret  = stats.get("total_return_pct", 0.0)
        ret_c = config.COLOR_UP if ret >= 0 else config.COLOR_DOWN
        ret_s = f"+{ret:.2f}" if ret >= 0 else f"{ret:.2f}"
        # 費率展示：模式 + 實際費率％
        fee_display = f"{fm} ({fr*100:.4f}%)" if fm not in ("Maker", "Taker") else f"{fm} ({fr*100:.2f}%)"
        cfg_lbl = QLabel(
            f"<b>資金:</b> {cap:,.0f} U &nbsp;|&nbsp; "
            f"<b>槓桿:</b> {lev}x &nbsp;|&nbsp; "
            f"<b>費率:</b> {fee_display} &nbsp;|&nbsp; "
            f"<b>損失上限:</b> {mlp*100:.1f}% &nbsp;|&nbsp; "
            f"<b>滑價:</b> {slip:.1f} bps &nbsp;|&nbsp; "
            f"<b>資金費率:</b> {fund:.4f}/8h &nbsp;|&nbsp; "
            f"<b>最終餘額:</b> {feq:,.2f} U "
            f"<span style='color:{ret_c}'>({ret_s}%)</span>"
        )
        cfg_lbl.setStyleSheet("font-size: 12px; padding: 4px 8px; color: #aaa;")
        layout.addWidget(cfg_lbl)

        # ── 回測區間 + Tick 覆蓋率 ──────────────────────────────────────
        from datetime import datetime as _dt2
        start_ms = stats.get("backtest_start_ms", 0)
        end_ms   = stats.get("backtest_end_ms", 0)
        tick_cov = stats.get("tick_coverage_pct")   # None = Bar 模式
        fallback = stats.get("fallback_bar_count", 0)

        def _fmt_ts(ms: int) -> str:
            if not ms:
                return "─"
            return _dt2.fromtimestamp(ms / 1000, tz=config.DISPLAY_TZ).strftime(
                "%Y-%m-%d %H:%M"
            )

        meta_parts = [f"<b>區間:</b> {_fmt_ts(start_ms)} ~ {_fmt_ts(end_ms)}"]
        if tick_cov is not None:
            meta_parts.append(f"<b>Tick 覆蓋:</b> {tick_cov:.1f}%")
            meta_parts.append(f"<b>無 Tick 棒:</b> {fallback}")
        meta_lbl = QLabel(" &nbsp;|&nbsp; ".join(meta_parts))
        meta_lbl.setStyleSheet("font-size: 12px; padding: 2px 8px; color: #80cbc4;")
        layout.addWidget(meta_lbl)


        _filter_style = (
            "QComboBox { background: #1e222d; color: #d1d4dc; border: 1px solid #363a45;"
            " border-radius: 3px; padding: 2px 6px; min-width: 90px; }"
            "QLabel { color: #aaa; font-size: 12px; }"
        )
        filter_w = QWidget()
        filter_w.setStyleSheet(_filter_style)
        filter_row = QHBoxLayout(filter_w)
        filter_row.setContentsMargins(4, 2, 4, 2)
        filter_row.addWidget(QLabel("市場時區:"))
        self._session_combo = QComboBox()
        for s in self.SESSIONS:
            self._session_combo.addItem(s)
        filter_row.addWidget(self._session_combo)
        filter_row.addSpacing(16)
        filter_row.addWidget(QLabel("月份:"))
        self._month_combo = QComboBox()
        self._populate_month_combo()
        filter_row.addWidget(self._month_combo)
        filter_row.addStretch()
        layout.addWidget(filter_w)

        # ── 摘要統計表（16 欄）────────────────────────────────────────
        self._sum_headers = [
            "策略", "交易數", "勝率", "PF", "淨利(USDT)", "手續費",
            "最大回撤", "最大連虧", "平均獲利", "平均虧損",
            "多單 PF", "空單 PF", "SL", "TP", "TS", "TD",
        ]
        self._sum_table = QTableWidget(1, len(self._sum_headers))
        self._sum_table.setHorizontalHeaderLabels(self._sum_headers)
        self._sum_table.verticalHeader().setVisible(False)
        self._sum_table.setMaximumHeight(58)
        self._sum_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._sum_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self._sum_table)

        self._warn_lbl = QLabel()
        self._warn_lbl.setStyleSheet("color:#f0c040; font-size:12px; padding:2px 8px;")
        self._warn_lbl.setVisible(False)
        layout.addWidget(self._warn_lbl)

        # ── 圖表快照按鈕 ─────────────────────────────────────────────
        self._snap_btn = QPushButton("📸 圖表快照")
        self._snap_btn.setToolTip("開啟所選交易的 K 棒快照視窗（雙擊表格列也可開啟）")
        self._snap_btn.setEnabled(bool(self._klines and self._full_trade_list))
        self._snap_btn.setStyleSheet(
            "QPushButton { background:#1e222d; color:#80cbc4;"
            " border:1px solid #26a69a; border-radius:3px; padding:3px 10px; }"
            "QPushButton:hover { background:#1a3a3a; }"
            "QPushButton:disabled { color:#444; border-color:#333; }"
        )
        self._snap_btn.clicked.connect(self._open_snapshot_selected)

        # ── 匯出按鈕 ─────────────────────────────────────────────────
        _exp_btn = QPushButton("⬇ 匯出 Excel")
        _exp_btn.setStyleSheet(
            "QPushButton { background:#1e3a1e; color:#26a69a; border:1px solid #26a69a;"
            " border-radius:3px; padding:3px 10px; }"
            "QPushButton:hover { background:#1a4a2a; }"
        )
        _exp_btn.clicked.connect(self._export_excel)
        _btn_row = QHBoxLayout()
        _btn_row.addWidget(self._snap_btn)
        _btn_row.addStretch()
        _btn_row.addWidget(_exp_btn)
        layout.addLayout(_btn_row)

        # ── 交易明細表 ────────────────────────────────────────────────
        self._trade_cols = [
            "#", "方向", "入場時間", "入場價", "出場類型",
            "出場價", "數量", "手續費", "資金費", "淨利(USDT)", "餘額", "Regime",
        ]
        self._trade_table = QTableWidget(0, len(self._trade_cols))
        self._trade_table.setHorizontalHeaderLabels(self._trade_cols)
        self._trade_table.verticalHeader().setVisible(False)
        self._trade_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._trade_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        self._trade_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self._trade_table.doubleClicked.connect(self._on_trade_double_clicked)
        layout.addWidget(self._trade_table)

        # ── 連接篩選信號 + 初次填充 ──────────────────────────────────
        self._session_combo.currentIndexChanged.connect(self._apply_filter)
        self._month_combo.currentIndexChanged.connect(self._apply_filter)
        self._apply_filter()

    # ── 過濾邏輯 ─────────────────────────────────────────────────────
    def _populate_month_combo(self) -> None:
        months = sorted(set(
            self._trade_month(t)
            for t in self._full_trade_list
            if not t.get("skipped") and self._trade_month(t)
        ))
        self._month_combo.addItem("全部月份")
        for m in months:
            self._month_combo.addItem(m)

    def _filtered_trades(self) -> list:
        trades = self._full_trade_list
        sess_key = self.SESSIONS.get(self._session_combo.currentText())
        if sess_key:
            trades = [t for t in trades if self._in_session(t, sess_key)]
        month = self._month_combo.currentText()
        if month != "全部月份":
            trades = [t for t in trades if self._trade_month(t) == month]
        return trades

    def _apply_filter(self) -> None:
        trades = self._filtered_trades()
        from backtest.engine import compute_subset_stats
        sub = compute_subset_stats(trades)
        sub["strategy_name"] = self._stats.get("strategy_name", "─")
        opens = self._stats.get("open_count", 0)
        self._refresh_summary(sub, opens)
        self._refresh_trades(trades)

    # ── 摘要表重新整理 ───────────────────────────────────────────────
    def _refresh_summary(self, s: dict, opens: int = 0) -> None:
        n    = s.get("trades", 0)
        net  = s.get("total_net_pnl", 0.0)
        net_s = f"+{net:,.2f}" if net >= 0 else f"{net:,.2f}"
        row_data = [
            s.get("strategy_name", "─"),
            str(n),
            f"{s.get('win_rate', 0.0):.1f}%",
            self._fmt_pf(s.get("profit_factor", 0.0)),
            net_s,
            f"{s.get('total_fees', 0.0):,.2f}",
            f"{s.get('max_drawdown_pct', 0.0):.2f}%",
            str(s.get("max_consec_loss", 0)),
            f"+{s.get('avg_win', 0.0):,.2f}" if s.get("avg_win", 0) > 0 else "─",
            f"-{s.get('avg_loss', 0.0):,.2f}" if s.get("avg_loss", 0) > 0 else "─",
            self._fmt_pf(s.get("long_profit_factor", 0.0)),
            self._fmt_pf(s.get("short_profit_factor", 0.0)),
            str(s.get("sl_count", 0)),
            str(s.get("tp_count", 0)),
            str(s.get("ts_count", 0)),
            str(s.get("td_count", 0)),
        ]
        for col, val in enumerate(row_data):
            item = QTableWidgetItem(val)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if col == 4:
                item.setForeground(QtGui.QColor(
                    config.COLOR_UP if net >= 0 else config.COLOR_DOWN
                ))
            self._sum_table.setItem(0, col, item)

        if opens:
            self._warn_lbl.setText(f"  ⚠ 未平倉: {opens} 筆")
            self._warn_lbl.setVisible(True)
        else:
            self._warn_lbl.setVisible(False)

    # ── 明細表重新整理 ───────────────────────────────────────────────
    def _refresh_trades(self, trades: list) -> None:
        from datetime import datetime
        active = [t for t in trades if not t.get("skipped")]
        self._trade_table.setRowCount(len(active))
        if not active:
            return

        _label_colors = {
            "SL": "#ef5350", "TP": "#26a69a",
            "TS": "#ff9800", "TD": "#9c27b0",
        }
        for i, t in enumerate(active):
            self._trade_table.setItem(i, 0, QTableWidgetItem(str(i + 1)))

            dir_txt = "做多" if t["dir"] == "long" else "做空"
            dir_item = QTableWidgetItem(dir_txt)
            dir_item.setForeground(
                QtGui.QColor(config.COLOR_UP if t["dir"] == "long" else config.COLOR_DOWN)
            )
            self._trade_table.setItem(i, 1, dir_item)

            # 入場時間（DISPLAY_TZ）
            ets = t.get("entry_time", 0)
            if ets:
                dt = datetime.fromtimestamp(ets / 1000, tz=config.DISPLAY_TZ)
                time_str = dt.strftime("%m-%d %H:%M")
            else:
                time_str = "─"
            self._trade_table.setItem(i, 2, QTableWidgetItem(time_str))

            self._trade_table.setItem(i, 3, QTableWidgetItem(f"{t['entry']:,.4f}"))

            exit_lbl = t.get("exit_label", "")
            lbl_item = QTableWidgetItem(exit_lbl)
            lbl_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if exit_lbl in _label_colors:
                lbl_item.setForeground(QtGui.QColor(_label_colors[exit_lbl]))
            self._trade_table.setItem(i, 4, lbl_item)

            self._trade_table.setItem(i, 5, QTableWidgetItem(f"{t['exit']:,.4f}"))
            self._trade_table.setItem(i, 6, QTableWidgetItem(f"{t.get('qty', 0):,.6f}"))
            self._trade_table.setItem(i, 7, QTableWidgetItem(f"{t.get('total_fee', 0):,.2f}"))
            self._trade_table.setItem(i, 8, QTableWidgetItem(f"{t.get('funding_cost', 0):,.2f}"))

            pv = t.get("net_pnl", 0.0)
            pnl_txt = f"+{pv:,.2f}" if pv >= 0 else f"{pv:,.2f}"
            pnl_item = QTableWidgetItem(pnl_txt)
            pnl_item.setForeground(
                QtGui.QColor(config.COLOR_UP if pv >= 0 else config.COLOR_DOWN)
            )
            self._trade_table.setItem(i, 9, pnl_item)
            self._trade_table.setItem(i, 10, QTableWidgetItem(f"{t.get('equity_after', 0):,.2f}"))

            _regime_colors = {
                "trend_up":   "#26a69a",
                "trend_down": "#ef5350",
                "range":      "#ff9800",
                "neutral":    "#9e9e9e",
            }
            _regime_labels = {
                "trend_up":   "↑ 趨勢",
                "trend_down": "↓ 趨勢",
                "range":      "◈ 盤整",
                "neutral":    "— 中性",
            }
            regime_val = t.get("regime", "")
            regime_item = QTableWidgetItem(_regime_labels.get(regime_val, regime_val or "─"))
            regime_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if regime_val in _regime_colors:
                regime_item.setForeground(QtGui.QColor(_regime_colors[regime_val]))
            self._trade_table.setItem(i, 11, regime_item)

    # ── Excel 匯出 ───────────────────────────────────────────────────
    def _export_excel(self) -> None:
        """將回測摘要與交易明細匯出為 Excel 檔案。"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(self, "缺少套件",
                                "請先安裝 openpyxl：\npip install openpyxl")
            return

        s = self._stats
        _strategy_raw = s.get("strategy_name", "backtest") or "backtest"
        _strategy_safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in _strategy_raw)
        from datetime import datetime as _dt2, timezone as _tz2
        def _fmt_date(ms: int) -> str:
            if not ms:
                return "unknown"
            return _dt2.fromtimestamp(ms / 1000, tz=_tz2.utc).strftime("%Y-%m-%d")
        _start_date = _fmt_date(s.get("backtest_start_ms", 0))
        _end_date   = _fmt_date(s.get("backtest_end_ms", 0))
        _default_name = f"{_strategy_safe}_{_start_date}_{_end_date}.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "匯出 Excel", _default_name,
            "Excel 檔案 (*.xlsx)"
        )
        if not path:
            return

        h_font   = Font(bold=True, color="FFFFFF")
        h_fill   = PatternFill("solid", fgColor="2962FF")
        s_fill   = PatternFill("solid", fgColor="1e222d")
        center   = Alignment(horizontal="center", vertical="center")

        wb = openpyxl.Workbook()

        # ── Sheet 1: 摘要 ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "摘要"

        param_heads = ["資金(U)", "槓桿", "費率模式", "費率%",
                       "損失上限%", "滑價(bps)", "資金費率/8h",
                       "最終餘額", "報酬率%",
                       "回測起始", "回測結束", "Tick覆蓋%", "無Tick棒數"]
        for col, h in enumerate(param_heads, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = h_font, h_fill, center

        param_vals = [
            s.get("initial_capital", 0),
            s.get("leverage", 0),
            s.get("fee_mode", ""),
            round(s.get("fee_rate", 0) * 100, 4),
            round(s.get("max_loss_pct", 0) * 100, 1),
            s.get("slippage_bps", 0.0),
            s.get("funding_rate", 0.0),
            round(s.get("final_equity", 0.0), 2),
            round(s.get("total_return_pct", 0.0), 2),
        ]
        # 回測區間與 Tick 統計
        from datetime import datetime as _dt3, timezone as _tz3
        def _fmt_ts_excel(ms: int) -> str:
            if not ms:
                return "─"
            return _dt3.fromtimestamp(ms / 1000, tz=_tz3.utc).strftime("%Y-%m-%d %H:%M")

        param_vals += [
            _fmt_ts_excel(s.get("backtest_start_ms", 0)),
            _fmt_ts_excel(s.get("backtest_end_ms", 0)),
            round(s["tick_coverage_pct"], 1) if s.get("tick_coverage_pct") is not None else "─",
            s.get("fallback_bar_count", 0),
        ]
        for col, val in enumerate(param_vals, 1):
            ws1.cell(row=2, column=col, value=val).alignment = center

        def _pf(v):
            return "∞" if v == float("inf") else round(v, 2)

        sum_heads = ["策略", "交易數", "勝率%", "PF",
                     "淨利(USDT)", "手續費",
                     "最大回撤%", "最大連虧",
                     "平均獲利", "平均虧損",
                     "多單PF", "空單PF",
                     "SL", "TP", "TS", "TD"]
        for col, h in enumerate(sum_heads, 1):
            c = ws1.cell(row=4, column=col, value=h)
            c.font, c.fill, c.alignment = h_font, s_fill, center

        sum_vals = [
            s.get("strategy_name", ""),
            s.get("trades", 0),
            round(s.get("win_rate", 0.0), 1),
            _pf(s.get("profit_factor", 0.0)),
            round(s.get("total_net_pnl", 0.0), 2),
            round(s.get("total_fees", 0.0), 2),
            round(s.get("max_drawdown_pct", 0.0), 2),
            s.get("max_consec_loss", 0),
            round(s.get("avg_win", 0.0), 2),
            round(s.get("avg_loss", 0.0), 2),
            _pf(s.get("long_profit_factor", 0.0)),
            _pf(s.get("short_profit_factor", 0.0)),
            s.get("sl_count", 0),
            s.get("tp_count", 0),
            s.get("ts_count", 0),
            s.get("td_count", 0),
        ]
        for col, val in enumerate(sum_vals, 1):
            cell = ws1.cell(row=5, column=col, value=val)
            cell.alignment = center
            if col == 5:
                net = s.get("total_net_pnl", 0.0)
                cell.font = Font(color="26A69A" if net >= 0 else "EF5350")

        for col in range(1, max(len(param_heads), len(sum_heads)) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 15

        # ── Sheet 2: 交易明細（當前篩選）─────────────────────────────
        ws2 = wb.create_sheet("交易明細")
        trade_heads = ["#", "方向", "入場時間", "入場價", "出場類型",
                       "出場價", "數量", "手續費",
                       "資金費", "淨利(USDT)", "餘額", "Regime"]
        trade_heads += [
            "side", "wick_type", "session_hour", "atr_percentile",
            "trend_regime", "entry_delay_bars", "k0_range_atr",
            "wick_volume_ratio", "zoom_delta_eff", "MAE", "MFE",
            "exit_label",
        ]
        for col, h in enumerate(trade_heads, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = h_font, h_fill, center

        from datetime import datetime as _dt
        filtered = self._filtered_trades()
        active = [t for t in filtered if not t.get("skipped")]
        for i, t in enumerate(active, 1):
            dir_txt = "做多" if t["dir"] == "long" else "做空"
            pv = t.get("net_pnl", 0.0)
            ets = t.get("entry_time", 0)
            time_str = _dt.fromtimestamp(ets / 1000, tz=config.DISPLAY_TZ).strftime(
                "%Y-%m-%d %H:%M") if ets else "─"
            row_vals = [
                i, dir_txt, time_str,
                round(t.get("entry", 0), 4),
                t.get("exit_label", ""),
                round(t.get("exit", 0), 4),
                round(t.get("qty", 0), 6),
                round(t.get("total_fee", 0), 2),
                round(t.get("funding_cost", 0), 2),
                round(pv, 2),
                round(t.get("equity_after", 0), 2),
                t.get("regime", "─"),
            ]
            row_vals += [
                t.get("side", ""),
                t.get("wick_type", ""),
                t.get("session_hour", ""),
                t.get("atr_percentile", ""),
                t.get("trend_regime", ""),
                t.get("entry_delay_bars", ""),
                t.get("k0_range_atr", ""),
                t.get("wick_volume_ratio", ""),
                t.get("zoom_delta_eff", ""),
                t.get("MAE", ""),
                t.get("MFE", ""),
                t.get("exit_label", ""),
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=i + 1, column=col, value=val)
                cell.alignment = center
                if col == 10:
                    cell.font = Font(color="26A69A" if pv >= 0 else "EF5350")

        for col in range(1, len(trade_heads) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 14

        wb.save(path)
        QMessageBox.information(self, "匯出成功", f"已儲存至:\n{path}")

    # ── 快照入口 ──────────────────────────────────────────────────────────────

    def _open_snapshot_at(self, trade_row_idx: int) -> None:
        """Build contexts once and open TradeSnapshotDialog at trade_row_idx."""
        if not self._klines:
            QMessageBox.information(
                self, "無法開啟快照",
                "此回測結果沒有可用的 K 棒快照資料，請重新執行新版回測。"
            )
            return

        from ui.trade_snapshot_dialog import TradeSnapshotDialog, _collect_contexts
        active_trades = [t for t in self._full_trade_list if not t.get("skipped")]
        contexts = []
        if self._signals:
            contexts = _collect_contexts(self._signals, active_trades, self._klines)
        if not contexts:
            contexts = _fallback_snapshot_contexts(active_trades, self._klines)
        if not contexts:
            QMessageBox.information(self, "無快照資料", "找不到可顯示的交易快照。")
            return

        # trade_row_idx 是過濾後表格的列號；contexts 是依完整清單建立的。
        # 先從過濾後清單取得目標 trade 物件，再在 contexts 中以 identity 比對找正確 index。
        filtered_active = [t for t in self._filtered_trades() if not t.get("skipped")]
        if trade_row_idx < len(filtered_active):
            target_trade = filtered_active[trade_row_idx]
            ctx_idx = next(
                (
                    i for i, c in enumerate(contexts)
                    if c["trade"] is target_trade or (
                        c["trade"].get("entry_time") == target_trade.get("entry_time")
                        and c["trade"].get("exit_time") == target_trade.get("exit_time")
                        and c["trade"].get("dir") == target_trade.get("dir")
                    )
                ),
                0,
            )
        else:
            ctx_idx = min(trade_row_idx, len(contexts) - 1)

        dlg = TradeSnapshotDialog(
            contexts, self._klines, self._tick_map,
            start_idx=ctx_idx, parent=self,
        )
        dlg.exec()

    def _open_snapshot_selected(self) -> None:
        """Open snapshot for the currently selected trade row (or first trade)."""
        rows = self._trade_table.selectionModel().selectedRows()
        row  = rows[0].row() if rows else 0
        self._open_snapshot_at(row)

    def _on_trade_double_clicked(self, index) -> None:
        self._open_snapshot_at(index.row())




class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        import config.base as _cfg_base
        self.setWindowTitle(f"{_cfg_base.APP_NAME} — Binance Futures")
        self.resize(1600, 900)

        # ── 狀態 ─────────────────────────────────────────────────────────────
        # 從設定載入上次的狀態
        self._symbol   = ui_settings.get("symbol", config.DEFAULT_SYMBOL)
        self._interval = ui_settings.get("interval", config.DEFAULT_INTERVAL)
        self._current_kline_open_time: int = 0
        self._last_price: float = 0.0
        self._loaded_klines: list = []   # 最近一次歷史 K 線，供 agg_history 使用
        self._kline_timestamps: List[int] = []  # kline open_time 序列，供 footprint x 軸對齊

        # ── 資料層 ────────────────────────────────────────────────────────────
        self._order_book = OrderBook()
        self._cvd_calc   = CvdCalculator()
        self._fp_builder = FootprintBuilder()

        self._ws_thread: Optional[WsWorkerThread] = None
        self._history_proc: Optional[HistoryProcessorThread] = None

        # ── 節流更新旗標 ──────────────────────────────────────────────────────
        self._dirty_cvd: bool = False
        self._dirty_fp: bool  = False
        self._last_trade_time: float = 0.0  # monotonic time of last aggTrade
        self._data_stale: bool = False       # 是否處於數據過期狀態
        # ── 策略測試狀態 ────────────────────────────────────────────
        self._strategy_engine: Optional[StrategyBase] = None
        self._strategy_realtime: bool = False
        self._strategy_signals: List[StrategySignal] = []
        self._jump_target_ms: Optional[int] = None   # 時間跳轉的待處理目標
        # ── UI ────────────────────────────────────────────────────────────────
        self._build_ui()
        self._build_toolbar()

        # 還原策略選擇
        last_strat = ui_settings.get("strategy_name")
        if last_strat and last_strat in STRATEGY_REGISTRY:
            self._strategy_combo.setCurrentText(last_strat)

        # ── Heatmap timer ─────────────────────────────────────────────────────
        self._heatmap_timer = QTimer(self)
        self._heatmap_timer.setInterval(config.HEATMAP_UPDATE_MS)
        self._heatmap_timer.timeout.connect(self._snapshot_heatmap)

        # ── 節流刷新 timer（150ms，避免每筆 trade 都重繪導致閃爍）─────────────
        self._flush_timer = QTimer(self)
        self._flush_timer.setInterval(150)
        self._flush_timer.timeout.connect(self._flush_updates)

        # ── 數據過期偵測 timer（每秒檢查一次，超過 5 秒無成交顯示警告）────────
        self._stale_timer = QTimer(self)
        self._stale_timer.setInterval(1000)
        self._stale_timer.timeout.connect(self._check_data_stale)

        # ── 啟動 ─────────────────────────────────────────────────────────────
        # K 線圖左滾觸發歷史載入（只連接一次，不隨 _start_stream 重建）
        self._kline_chart.need_more_history.connect(self._on_need_more_history)

        # ── 十字線同步 ────────────────────────────────────────────────────────
        self._crosshair_charts = [
            self._kline_chart, self._fp_chart, self._cvd_chart,
        ]
        for chart in self._crosshair_charts:
            chart.crosshair_moved.connect(self._sync_crosshair)
            chart.crosshair_left.connect(self._hide_all_crosshairs)

        self._start_stream()

    # ══════════════════════════════════════════════════════════════
    # UI 建構
    # ══════════════════════════════════════════════════════════════

    def _build_toolbar(self) -> None:
        # 嵌入式控制列（置於即時看盤分頁頂部）
        ctrl = QHBoxLayout(self._live_ctrl_bar)
        ctrl.setContentsMargins(6, 2, 6, 2)
        ctrl.setSpacing(4)

        _lbl_style = "color:#aaa; font-size:11px;"
        _sep_style  = "background:#2a2e39;"
        _btn_style  = (
            "QPushButton { background:#1e222d; color:#d1d4dc; border:1px solid #2a2e39;"
            " border-radius:3px; padding:1px 6px; font-size:11px; }"
            "QPushButton:checked { background:#2962ff; color:#fff; }"
            "QPushButton:hover   { background:#2a2e39; }"
        )

        def _sep() -> QFrame:
            f = QFrame()
            f.setFrameShape(QFrame.Shape.VLine)
            f.setFixedWidth(1)
            f.setStyleSheet(_sep_style)
            return f

        # Symbol
        _l = QLabel("交易對")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)
        self._sym_combo = QComboBox()
        self._sym_combo.addItems(config.SYMBOLS)
        self._sym_combo.setCurrentText(self._symbol)
        self._sym_combo.currentTextChanged.connect(self._on_symbol_changed)
        ctrl.addWidget(self._sym_combo)

        ctrl.addWidget(_sep())

        # Interval
        _l = QLabel("週期")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)
        self._iv_combo = QComboBox()
        self._iv_combo.addItems(config.INTERVALS)
        self._iv_combo.setCurrentText(self._interval)
        self._iv_combo.currentTextChanged.connect(self._on_interval_changed)
        ctrl.addWidget(self._iv_combo)

        ctrl.addWidget(_sep())

        # Footprint mode
        _l = QLabel("Footprint")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)
        self._fp_combo = QComboBox()
        self._fp_combo.addItems(config.FOOTPRINT_MODES)
        self._fp_combo.currentTextChanged.connect(self._on_fp_mode_changed)
        ctrl.addWidget(self._fp_combo)

        ctrl.addWidget(_sep())

        # Tick 聚合倍數
        _l = QLabel("Tick")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)
        self._tick_combo = QComboBox()
        for m in config.TICK_MULTIPLIERS:
            self._tick_combo.addItem(f"{m}x")
        self._tick_combo.setCurrentIndex(
            config.TICK_MULTIPLIERS.index(config.DEFAULT_TICK_MULTIPLIER)
        )
        self._tick_combo.currentIndexChanged.connect(self._on_tick_multiplier_changed)
        ctrl.addWidget(self._tick_combo)

        ctrl.addWidget(_sep())

        # Log scale 切換按鈕
        self._log_btn = QPushButton("Log")
        self._log_btn.setCheckable(True)
        self._log_btn.setFixedWidth(40)
        self._log_btn.setToolTip("切換 K 線 Y 軸：線性 ↔ 對數")
        self._log_btn.setStyleSheet(_btn_style)
        self._log_btn.toggled.connect(self._on_log_toggled)
        ctrl.addWidget(self._log_btn)

        ctrl.addWidget(_sep())

        # ── 即時策略標注 ────────────────────────────────────────────────────
        _l = QLabel("策略")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)
        self._strategy_combo = QComboBox()
        self._strategy_combo.addItem("── 無 ──")
        self._strategy_combo.addItems(list(STRATEGY_REGISTRY.keys()))
        self._strategy_combo.currentTextChanged.connect(self._on_strategy_changed)
        ctrl.addWidget(self._strategy_combo)

        ctrl.addWidget(_sep())

        self._rt_btn = QPushButton("⚡ 即時")
        self._rt_btn.setCheckable(True)
        self._rt_btn.setToolTip("開啟後，每根 K 棒收盤自動標註")
        self._rt_btn.setStyleSheet(_btn_style)
        self._rt_btn.toggled.connect(self._on_realtime_toggled)
        ctrl.addWidget(self._rt_btn)

        self._clear_btn = QPushButton("✕ 清除")
        self._clear_btn.setToolTip("清除標記與統計")
        self._clear_btn.setStyleSheet(_btn_style)
        self._clear_btn.clicked.connect(self._on_clear_strategy)
        ctrl.addWidget(self._clear_btn)

        self._strategy_stats_lbl = QLabel()
        self._strategy_stats_lbl.setStyleSheet(
            "color:#f0c040; font-size:11px; padding-left:6px;"
        )
        self._strategy_stats_lbl.setVisible(False)
        ctrl.addWidget(self._strategy_stats_lbl)

        ctrl.addWidget(_sep())

        self._status_lbl = QLabel("初始化中 …")
        self._status_lbl.setStyleSheet("color: #aaa; font-size: 11px;")
        ctrl.addWidget(self._status_lbl)

        _l = QLabel("Data")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)

        self._data_root_lbl = QLabel()
        self._data_root_lbl.setStyleSheet("color:#80cbc4; font-size:11px;")
        self._data_root_lbl.setMinimumWidth(130)
        ctrl.addWidget(self._data_root_lbl)

        self._data_root_btn = QPushButton("...")
        self._data_root_btn.setFixedWidth(28)
        self._data_root_btn.setToolTip("Select OrderFlow data root")
        self._data_root_btn.setStyleSheet(_btn_style)
        self._data_root_btn.clicked.connect(self._on_choose_data_root)
        ctrl.addWidget(self._data_root_btn)

        self._refresh_data_root_status()

        self._price_lbl = QLabel("─")
        self._price_lbl.setStyleSheet(
            "color: #d1d4dc; font-size: 13px; font-weight: bold; padding-left: 10px;"
        )
        ctrl.addWidget(self._price_lbl)

        # ── 時間跳轉（置右）────────────────────────────────────────────────
        ctrl.addStretch(1)

        ctrl.addWidget(_sep())

        _l = QLabel("跳轉至")
        _l.setStyleSheet(_lbl_style)
        ctrl.addWidget(_l)

        self._jump_dt_edit = QDateTimeEdit()
        self._jump_dt_edit.setDisplayFormat("yyyy-MM-dd HH:mm")
        self._jump_dt_edit.setTimeSpec(Qt.TimeSpec.UTC)
        self._jump_dt_edit.setDateTime(QDateTime.currentDateTimeUtc())
        self._jump_dt_edit.setToolTip("指定要跳轉的 K 棒時間（UTC）")
        self._jump_dt_edit.setFixedWidth(138)
        ctrl.addWidget(self._jump_dt_edit)

        self._jump_btn = QPushButton("跳轉")
        self._jump_btn.setToolTip("跳轉到指定時間的 K 棒")
        self._jump_btn.setStyleSheet(_btn_style)
        self._jump_btn.clicked.connect(self._on_jump_to_time)
        ctrl.addWidget(self._jump_btn)

        if hasattr(self, "_top_tabs"):
            self._top_tabs.currentChanged.connect(self._on_top_tab_changed)

    def _build_ui(self) -> None:
        # ── 左欄（OB + Heatmap）─────────────────────────────────────────────
        self._ob_widget  = OrderBookWidget()
        self._heatmap    = HeatmapWidget()

        left_widget = QWidget()
        left_lay = QVBoxLayout(left_widget)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(2)

        ob_label = QLabel("Order Book")
        ob_label.setStyleSheet("color:#aaa; font-size:10px; padding:2px 4px;")
        left_lay.addWidget(ob_label)
        left_lay.addWidget(self._ob_widget, 3)   # 佔 3/5

        hm_label = QLabel("OB Heatmap")
        hm_label.setStyleSheet("color:#aaa; font-size:10px; padding:2px 4px;")
        left_lay.addWidget(hm_label)
        left_lay.addWidget(self._heatmap, 2)      # 佔 2/5

        left_widget.setMinimumWidth(200)
        left_widget.setMaximumWidth(280)

        # ── 右欄（K線/Footprint tab + CVD）────────────────────────────────
        self._kline_chart = KlineChart()
        self._fp_chart    = FootprintChart()
        self._cvd_chart   = CvdChart()
        self._stats_panel = StatsPanel()

        # K線 / Footprint tab
        self._chart_tabs = QTabWidget()
        self._chart_tabs.setTabPosition(QTabWidget.TabPosition.North)
        self._chart_tabs.setStyleSheet(
            "QTabBar::tab { padding: 4px 14px; }"
        )
        self._chart_tabs.addTab(self._kline_chart, "K 線")
        self._chart_tabs.addTab(self._fp_chart,    "Footprint")
        self._capacity_tab = CapacityTab()
        self._chart_tabs.addTab(self._capacity_tab, "容量分析")

        # 連結 CVD / Stats 的 x 軸到 K 線
        self._cvd_chart.link_x(self._kline_chart.get_plot_item())
        self._stats_panel.link_x(self._kline_chart.get_plot_item())

        # Footprint 也與 K 線共用 x 軸座標空間
        self._fp_chart.get_plot_item().setXLink(self._kline_chart.get_plot_item())

        # StatsPanel 固定高度
        self._stats_panel.setMaximumHeight(82)
        self._stats_panel.setMinimumHeight(60)

        right_splitter = QSplitter(Qt.Orientation.Vertical)
        right_splitter.addWidget(self._chart_tabs)
        right_splitter.addWidget(self._cvd_chart)
        right_splitter.addWidget(self._stats_panel)
        right_splitter.setSizes([600, 160, 82])

        # ── 主橫向分割（即時看盤頁面）────────────────────────────────────────────
        main_splitter = QSplitter(Qt.Orientation.Horizontal)
        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_splitter)
        main_splitter.setSizes([230, 1370])

        # ── 即時看盤分頁容器（含嵌入控制列）────────────────────────────────────
        live_tab_widget = QWidget()
        live_tab_layout = QVBoxLayout(live_tab_widget)
        live_tab_layout.setContentsMargins(0, 0, 0, 0)
        live_tab_layout.setSpacing(0)

        self._live_ctrl_bar = QFrame()
        self._live_ctrl_bar.setStyleSheet(
            "QFrame { background: #1a1f2e; border-bottom: 1px solid #2a2e39; }"
        )
        self._live_ctrl_bar.setFixedHeight(36)
        live_tab_layout.addWidget(self._live_ctrl_bar)
        live_tab_layout.addWidget(main_splitter, stretch=1)

        # ── 頂層頁籤：Tab 0=回測分析  Tab 1=即時看盤 ─────────────────────────
        from ui.backtest_dashboard import BacktestDashboard
        from ui.research_lab import ResearchLab
        from ui.pipeline_studio import PipelineStudio
        self._backtest_dashboard = BacktestDashboard()
        self._research_lab = ResearchLab()
        self._pipeline_studio = PipelineStudio()

        self.setDockOptions(
            QMainWindow.DockOption.AllowNestedDocks
            | QMainWindow.DockOption.AllowTabbedDocks
            | QMainWindow.DockOption.AnimatedDocks
        )
        self.setTabPosition(
            Qt.DockWidgetArea.AllDockWidgetAreas,
            QTabWidget.TabPosition.North,
        )
        self.setStyleSheet(
            self.styleSheet()
            + """
            QMainWindow::separator {
                background: #252b3a;
                width: 4px;
                height: 4px;
            }
            QDockWidget {
                color: #d1d4dc;
                font-size: 12px;
            }
            QDockWidget::title {
                background: #151b28;
                border: 1px solid #263245;
                border-bottom: 0;
                padding: 7px 10px;
                text-align: left;
            }
            QTabBar::tab {
                background: #151b28;
                color: #8f96a8;
                border: 1px solid #263245;
                border-bottom: 0;
                padding: 7px 18px;
                min-width: 128px;
            }
            QTabBar::tab:selected {
                background: #20283a;
                color: #f2f5f9;
                border-top: 2px solid #26a69a;
            }
            QTabBar::tab:hover {
                color: #d1d4dc;
                background: #1d2535;
            }
            """
        )

        self._page_docks: dict[str, QDockWidget] = {}
        dock_specs = [
            ("backtest", "回測分析", self._backtest_dashboard),
            ("live", "即時看盤", live_tab_widget),
            ("research", "Research Lab", self._research_lab),
            ("pipeline", "Pipeline 設計室", self._pipeline_studio),
        ]
        first_dock: QDockWidget | None = None
        for key, title, widget in dock_specs:
            dock = self._create_page_dock(key, title, widget)
            self.addDockWidget(Qt.DockWidgetArea.LeftDockWidgetArea, dock)
            if first_dock is None:
                first_dock = dock
            else:
                self.tabifyDockWidget(first_dock, dock)
            self._page_docks[key] = dock
        if first_dock is not None:
            first_dock.raise_()

    def _on_top_tab_changed(self, index: int) -> None:
        pass  # 各分頁控制列已嵌入於分頁內，無需切換顯示

    # ══════════════════════════════════════════════════════════════
    # WebSocket 管理
    # ══════════════════════════════════════════════════════════════

    def _create_page_dock(self, key: str, title: str, widget: QWidget) -> QDockWidget:
        dock = QDockWidget(title, self)
        dock.setObjectName(f"page_dock_{key}")
        dock.setAllowedAreas(Qt.DockWidgetArea.AllDockWidgetAreas)
        dock.setFeatures(
            QDockWidget.DockWidgetFeature.DockWidgetMovable
            | QDockWidget.DockWidgetFeature.DockWidgetFloatable
        )
        dock.setWidget(widget)
        dock.setMinimumSize(420, 280)
        dock.topLevelChanged.connect(
            lambda floating, d=dock: self._on_page_dock_floating_changed(d, floating)
        )
        return dock

    def _on_page_dock_floating_changed(self, dock: QDockWidget, floating: bool) -> None:
        if floating and (dock.width() < 900 or dock.height() < 560):
            dock.resize(1200, 720)

    def _refresh_data_root_status(self) -> None:
        root = data_paths.data_root()
        ok, message = data_paths.validate_data_root()
        self._data_root_lbl.setText(root.name or str(root))
        self._data_root_lbl.setToolTip(f"{root}\n{message}")
        self._data_root_lbl.setStyleSheet(
            "color:#80cbc4; font-size:11px;" if ok else "color:#f0c040; font-size:11px;"
        )

    def _on_choose_data_root(self) -> None:
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select OrderFlow data root",
            str(data_paths.data_root()),
        )
        if not folder:
            return
        root = data_paths.ensure_data_root_layout(folder)
        data_paths.set_data_root_override(root)
        ui_settings.set("data_root", str(root))
        self._refresh_data_root_status()
        if hasattr(self, "_backtest_dashboard"):
            refresh = getattr(self._backtest_dashboard, "refresh_data_root", None)
            if callable(refresh):
                refresh()
        self._status_lbl.setText(f"Data root: {root}")

    def _start_stream(self) -> None:
        # 停止舊執行緒
        if self._history_proc and self._history_proc.isRunning():
            self._history_proc.quit()
            self._history_proc.wait(2000)
            self._history_proc = None
        if self._ws_thread:
            self._ws_thread.stop()
            self._ws_thread.wait(4000)
            self._ws_thread = None

        # 重置資料層
        self._order_book.reset()
        self._cvd_calc.reset()
        self._fp_builder.reset(
            tick_size=config.TICK_SIZES.get(self._symbol, 1.0)
        )
        self._fp_chart.set_tick_size(
            config.TICK_SIZES.get(self._symbol, 1.0)
        )
        self._current_kline_open_time = 0
        self._last_price = 0.0
        self._last_trade_time = 0.0
        self._data_stale = False
        self._loaded_klines = []
        self._kline_timestamps = []

        # 重置 UI
        self._kline_chart.set_history([])
        self._heatmap.reset()
        self._cvd_chart.update_cvd([])
        self._stats_panel.update_data([], [])
        self._fp_chart.reset_auto_range()

        # 重置策略狀態（清標記、統計；保留選取的策略 engine）
        self._on_clear_strategy()
        self._strategy_realtime = False
        self._rt_btn.setChecked(False)

        # 建立新執行緒
        self._ws_thread = WsWorkerThread(self._symbol, self._interval)
        self._ws_thread.trade_signal.connect(self._on_trade)
        self._ws_thread.kline_signal.connect(self._on_kline)
        self._ws_thread.depth_signal.connect(self._on_depth)
        self._ws_thread.ob_snapshot_signal.connect(self._on_ob_snapshot)
        self._ws_thread.history_signal.connect(self._on_history)
        self._ws_thread.agg_history_signal.connect(self._on_agg_history)
        self._ws_thread.more_history_signal.connect(self._on_more_history)
        self._ws_thread.more_agg_history_signal.connect(self._on_more_agg_history)
        self._ws_thread.exchange_info_signal.connect(self._on_exchange_info)
        self._ws_thread.status_signal.connect(self._on_status)
        self._ws_thread.cache_ready_signal.connect(self._on_cache_ready)
        self._ws_thread.start()

        self._heatmap_timer.start()
        self._flush_timer.start()
        self._stale_timer.start()

    # ══════════════════════════════════════════════════════════════
    # Signal handlers（主執行緒）
    # ══════════════════════════════════════════════════════════════

    def _on_trade(self, data: dict) -> None:
        trade = Trade(
            symbol=data["s"],
            price=float(data["p"]),
            qty=float(data["q"]),
            is_buyer_maker=bool(data["m"]),
            trade_time=int(data["T"]),
        )
        self._last_price = trade.price
        self._last_trade_time = time.monotonic()

        # 若之前處於斷線/延遲狀態，恢復正常
        if self._data_stale:
            self._data_stale = False
            self._status_lbl.setStyleSheet("color: #aaa; font-size: 11px;")
            self._status_lbl.setText(f"已連線：{self._symbol} {self._interval}")

        # CVD — 只累加數值，不立即重繪
        self._cvd_calc.update(trade)
        self._dirty_cvd = True

        # Footprint — 只更新資料結構，不立即重繪
        self._fp_builder.update_trade(trade, open_time=self._current_kline_open_time)
        self._dirty_fp = True

        # Heatmap 成交點（低成本，不節流）
        self._heatmap.add_trade(
            trade.price, trade.qty, not trade.is_buyer_maker
        )

        # 更新價格標籤（低成本）
        color = "#26a69a" if not trade.is_buyer_maker else "#ef5350"
        self._price_lbl.setStyleSheet(
            f"color:{color}; font-size:14px; font-weight:bold; padding-left:16px;"
        )
        self._price_lbl.setText(f"{self._symbol}  {trade.price:,.4f}")

    def _on_kline(self, data: dict) -> None:
        k_raw = data["k"]
        kline = Kline.from_ws(k_raw)

        # ── 同步 _loaded_klines（即時策略的前提條件）──────────────────────
        if self._loaded_klines:
            if self._loaded_klines[-1].open_time == kline.open_time:
                self._loaded_klines[-1] = kline
            else:
                self._loaded_klines.append(kline)
                # 內存管理：限制最大長度
                if len(self._loaded_klines) > config.KLINE_MAX_LOADED:
                    self._loaded_klines = self._loaded_klines[-config.KLINE_MAX_LOADED:]

        # 通知 CVD 新 K 棒開始
        if kline.open_time != self._current_kline_open_time:
            self._cvd_calc.on_new_candle(kline.open_time)
            self._current_kline_open_time = kline.open_time
            self._fp_builder.set_current_open_time(kline.open_time)
            # 更新 kline timestamps 供 footprint 對齊（並上限 KLINE_MAX_LOADED 防長時間記憶體增長）
            if not self._kline_timestamps or self._kline_timestamps[-1] != kline.open_time:
                self._kline_timestamps.append(kline.open_time)
                if len(self._kline_timestamps) > config.KLINE_MAX_LOADED:
                    self._kline_timestamps = self._kline_timestamps[-config.KLINE_MAX_LOADED:]
                self._fp_chart.set_kline_timestamps(self._kline_timestamps)
                self._fp_builder.set_kline_open_times(self._kline_timestamps)

        # 更新 Footprint OHLCV
        self._fp_builder.update_kline(kline)

        # 更新 K 線圖
        self._kline_chart.update_candle(kline)

        # ── 即時策略標註（K 棒收盤後觸發）──────────────────────────
        if (
            kline.is_closed
            and self._strategy_realtime
            and self._strategy_engine is not None
            and self._loaded_klines
        ):
            sig = self._strategy_engine.on_kline(kline, self._loaded_klines)
            if sig is not None:
                self._strategy_signals.append(sig)
                # 上限 2000 筆，防止長時間燒機時 set_strategy_markers O(N) 惡化
                if len(self._strategy_signals) > 2000:
                    self._strategy_signals = self._strategy_signals[-2000:]
                self._kline_chart.set_strategy_markers(self._strategy_signals)
                self._update_strategy_stats()

    def _on_depth(self, data: dict) -> None:
        needs_resync = self._order_book.apply_diff(data)
        if needs_resync and self._ws_thread:
            self._ws_thread.request_resync()
            return

        bids = self._order_book.get_bids(config.OB_DISPLAY_LEVELS)
        asks = self._order_book.get_asks(config.OB_DISPLAY_LEVELS)
        self._ob_widget.update_ob(bids, asks, self._last_price)

    def _on_ob_snapshot(self, data: dict) -> None:
        if not data:
            return
        self._order_book.init_snapshot(data)
        logger.debug("OB snapshot received")

    def _on_exchange_info(self, tick_map: dict) -> None:
        """從 exchangeInfo 動態更新交易所原始 tick，並為未知幣種推算顯示用 tick。"""
        config.EXCHANGE_TICK_SIZES.update(tick_map)

        # 僅對 TICK_SIZES 中尚未定義的幣種，自動推算合理的顯示用 tick
        for symbol, raw_tick in tick_map.items():
            if symbol not in config.TICK_SIZES:
                # 啟發式：取交易所 tick 的 100 倍作為初始顯示 tick
                # （多數幣種的原始 tick 遠小於可讀的分桶大小）
                config.TICK_SIZES[symbol] = raw_tick * 100
                logger.info(
                    "Auto-computed display tick for %s: exchange=%.10g → display=%.10g",
                    symbol, raw_tick, raw_tick * 100,
                )

        # 用目前選中的聚合倍數重新計算實際 tick size
        base_tick = config.TICK_SIZES.get(self._symbol, 1.0)
        multiplier = config.TICK_MULTIPLIERS[self._tick_combo.currentIndex()]
        effective_tick = base_tick * multiplier
        self._fp_builder.reset(tick_size=effective_tick)
        self._fp_chart.set_tick_size(effective_tick)
        logger.info(
            "tickSize for %s: display_base=%.10g, multiplier=%dx, effective=%.10g",
            self._symbol, base_tick, multiplier, effective_tick,
        )

    def _on_history(self, rows: list) -> None:
        """從 REST 取得的歷史 K 線（list of list）。"""
        from core.data_types import Kline as _Kline
        klines = [
            _Kline.from_rest(self._symbol, self._interval, row)
            for row in rows
        ]
        self._loaded_klines = klines
        self._kline_chart.set_history(klines)

        if klines:
            # 傳遞 kline timestamps 給 footprint chart 供 x 軸對齊
            self._kline_timestamps = [k.open_time for k in klines]
            self._fp_chart.set_kline_timestamps(self._kline_timestamps)
            self._fp_builder.set_kline_open_times(self._kline_timestamps)

            # ── CVD: 從歷史 K 線的 taker_buy_volume 計算真正 CVD ──
            self._cvd_calc.seed_history(klines[:-1])
            self._cvd_calc.on_new_candle(klines[-1].open_time)
            self._current_kline_open_time = klines[-1].open_time
            self._fp_builder.set_current_open_time(klines[-1].open_time)

            # 立即將歷史 CVD 曲線渲染到畫面上
            ot_map = self._kline_chart.get_open_time_index_map()
            self._cvd_chart.update_cvd(self._cvd_calc.get_series(), ot_map)

    def _on_agg_history(self, payload_list: list) -> None:
        """
        收到歷史 aggTrades → 啟動背景執行緒處理，釋放主執行緒。
        payload_list[0] = {
            'trades': [...aggTrade dicts...],
            'klines': [(open_time_ms, close_time_ms), ...]
        }
        """
        if not payload_list:
            return
        payload = payload_list[0]
        if not payload.get("trades"):
            return

        # 若前一次處理尚未完成，先等它結束再取代
        if self._history_proc and self._history_proc.isRunning():
            self._history_proc.quit()
            self._history_proc.wait(3000)

        tick_size = config.TICK_SIZES.get(self._symbol, 1.0)
        history_klines = self._loaded_klines[-(config.FOOTPRINT_HISTORY_CANDLES):]

        self._history_proc = HistoryProcessorThread(
            payload=payload,
            tick_size=tick_size,
            history_klines=history_klines,
            parent=self,
        )
        self._history_proc.result_signal.connect(self._on_history_processed)
        self._history_proc.status_signal.connect(self._on_status)
        self._history_proc.start()

    def _on_need_more_history(self, oldest_open_time_ms: int) -> None:
        """K 線圖滚到最左端，請求更早的歷史資料。"""
        if self._ws_thread:
            self._ws_thread.request_more_history(oldest_open_time_ms)

    def _on_more_history(self, rows: list) -> None:
        """WsWorkerThread 完成更早歷史 K 線載入後回調。"""
        self._kline_chart.set_loading_more(False)
        if not rows:
            return

        from core.data_types import Kline as _Kline
        new_klines = [
            _Kline.from_rest(self._symbol, self._interval, row)
            for row in rows
        ]
        # 去除與現有資料重疊的部分
        if self._loaded_klines:
            cutoff = self._loaded_klines[0].open_time
            new_klines = [k for k in new_klines if k.open_time < cutoff]
        if not new_klines:
            return

        # 將新 K 棒插入歷史最前端
        self._loaded_klines = new_klines + self._loaded_klines
        # 內存管理：限制最大長度
        if len(self._loaded_klines) > config.KLINE_MAX_LOADED:
            self._loaded_klines = self._loaded_klines[-config.KLINE_MAX_LOADED:]
        self._kline_timestamps = [k.open_time for k in self._loaded_klines]
        self._fp_chart.set_kline_timestamps(self._kline_timestamps)
        self._fp_builder.set_kline_open_times(self._kline_timestamps)

        # Footprint 小圖已連結 x 軸，更新 timestamps 後重繪現有資料
        fp_candles = self._fp_builder.get_candles()
        if fp_candles:
            self._fp_chart.update_candles(fp_candles)

        # K 線圖前插新 K 棒並平移視圖
        self._kline_chart.prepend_history(new_klines)

        # 如果目前有策略標記，前插後 x 索引已整體右移，需重新渲染
        if self._strategy_signals:
            self._kline_chart.set_strategy_markers(self._strategy_signals)

        # 時間跳轉：若目標已在載入範圍內則執行跳轉，否則繼續載入更早的歷史
        if self._jump_target_ms is not None:
            oldest_ms = self._loaded_klines[0].open_time if self._loaded_klines else 0
            if oldest_ms <= self._jump_target_ms:
                self._kline_chart.scroll_to_time(self._jump_target_ms)
                self._jump_target_ms = None
            else:
                if self._ws_thread:
                    self._ws_thread.request_more_history(oldest_ms)

    def _on_more_agg_history(self, payload_list: list) -> None:
        """WsWorkerThread 完成更早 aggTrades 載入後回調，連同處理 Footprint。"""
        self._on_agg_history(payload_list)

    def _on_history_processed(self, candles: list) -> None:
        """背景執行緒完成後，將 Footprint K 棒合併進主資料層並更新 UI。"""
        if not candles:
            return

        # 將歷史 candles 寫入主 fp_builder（只覆蓋尚未有資料的 K 棒）
        for candle in candles:
            ot = candle.open_time
            if ot not in self._fp_builder._candles:
                self._fp_builder._candles[ot] = candle
            else:
                # 合併：保留實時累積的 levels，補齊 OHLCV
                live = self._fp_builder._candles[ot]
                if live.open == 0.0:
                    live.open   = candle.open
                    live.high   = candle.high
                    live.low    = candle.low
                    live.close  = candle.close
                    live.volume = candle.volume
                for price, lv in candle.levels.items():
                    if price not in live.levels:
                        live.levels[price] = lv

        fp_candles = self._fp_builder.get_candles()
        ot_map = self._kline_chart.get_open_time_index_map()
        self._fp_chart.update_candles(fp_candles, ot_map)
        self._stats_panel.update_data(fp_candles, self._cvd_calc.get_series(), ot_map)
        self._status_lbl.setText(f"已連線：{self._symbol} {self._interval}")

    def _flush_updates(self) -> None:
        """節流刷新：每 150ms 最多重繪一次 CVD / Footprint / Stats。"""
        if self._dirty_cvd or self._dirty_fp:
            cvd_series = self._cvd_calc.get_series()
            fp_candles = self._fp_builder.get_candles()
            # 統一使用 KlineChart 的当前索引映射，避免 kline 截斷後各圖表 x 坐標漂移
            ot_map = self._kline_chart.get_open_time_index_map()

            if self._dirty_cvd:
                self._cvd_chart.update_cvd(cvd_series, ot_map)
                self._dirty_cvd = False

            if self._dirty_fp:
                self._fp_chart.update_candles(fp_candles, ot_map)
                self._dirty_fp = False

            self._stats_panel.update_data(fp_candles, cvd_series, ot_map)

    def _on_status(self, msg: str) -> None:
        self._status_lbl.setText(msg)

    def _check_data_stale(self) -> None:
        """每秒檢查：若超過 5 秒未收到 aggTrade，標記數據過期。"""
        if self._last_trade_time <= 0:
            return
        elapsed = time.monotonic() - self._last_trade_time
        if elapsed > 5.0 and not self._data_stale:
            self._data_stale = True
            self._status_lbl.setStyleSheet(
                "color: #ff4444; font-size: 11px; font-weight: bold;"
            )
            self._status_lbl.setText(
                f"⚠ 數據延遲 ({elapsed:.0f}s 無成交) — 請留意行情可能不準確"
            )
        elif self._data_stale and elapsed > 5.0:
            # 持續更新延遲秒數
            self._status_lbl.setText(
                f"⚠ 數據延遲 ({elapsed:.0f}s 無成交) — 請留意行情可能不準確"
            )

    # ══════════════════════════════════════════════════════════════
    # Heatmap timer
    # ══════════════════════════════════════════════════════════════

    def _snapshot_heatmap(self) -> None:
        if not self._order_book.is_initialized:
            return
        mid = self._order_book.mid_price()
        if mid <= 0:
            mid = self._last_price
        if mid <= 0:
            return
        bids = self._order_book.get_bids(200)
        asks = self._order_book.get_asks(200)
        self._heatmap.add_snapshot(bids, asks, mid)

    # ══════════════════════════════════════════════════════════════
    # 工具列事件
    # ══════════════════════════════════════════════════════════════

    def _on_symbol_changed(self, sym: str) -> None:
        if sym == self._symbol:
            return
        self._symbol = sym
        ui_settings.set("symbol", sym)
        self._start_stream()

    def _on_interval_changed(self, iv: str) -> None:
        if iv == self._interval:
            return
        self._interval = iv
        ui_settings.set("interval", iv)
        self._start_stream()

    def _on_fp_mode_changed(self, mode: str) -> None:
        self._fp_chart.set_mode(mode)
        ui_settings.set("fp_mode", mode)

    def _on_tick_multiplier_changed(self, index: int) -> None:
        """切換價格聚合倍數，重新設定 tick size 並清空 Footprint。"""
        multiplier = config.TICK_MULTIPLIERS[index]
        ui_settings.set("tick_multiplier_index", index)
        base_tick = config.TICK_SIZES.get(self._symbol, 1.0)
        effective_tick = base_tick * multiplier
        self._fp_builder.reset(tick_size=effective_tick)
        self._fp_chart.set_tick_size(effective_tick)
        if self._kline_timestamps:
            self._fp_builder.set_kline_open_times(self._kline_timestamps)
        logger.info("Tick multiplier changed to %dx (tick=%.10g)", multiplier, effective_tick)

    def _on_log_toggled(self, checked: bool) -> None:
        """Log scale 按鈕切換。"""
        self._kline_chart.toggle_log_scale()
        lbl = "Log ✓" if checked else "Log"
        self._log_btn.setText(lbl)
        ui_settings.set("log_scale", checked)

    # ── 策略事件 ──────────────────────────────────────────────────────────────

    def _on_strategy_changed(self, name: str) -> None:
        """下拉選單切換策略 engine。"""
        ui_settings.set("strategy_name", name)
        if name == "── 無 ──":
            self._strategy_engine = None
        else:
            cls = STRATEGY_REGISTRY.get(name)
            self._strategy_engine = cls() if cls else None
    def _on_cache_ready(self, count: int) -> None:
        """K 線快取下載完成後的回調。"""
        if count:
            info = kline_cache.info(self._symbol, self._interval)
            if info:
                self._status_lbl.setText(
                    f"快取已儲存：{info['count']:,} 根 K 棒 ({info['size_mb']:.1f} MB)"
                )
            else:
                self._status_lbl.setText(f"快取已儲存：{count:,} 根")
        else:
            self._status_lbl.setText("快取下載失敗，請稍後再試")

    # ─────────────────────────────────────────────────────────────────────────

    def _on_jump_to_time(self) -> None:
        """跳轉到指定時間的 K 棒。若目標時間早於已載入資料則先載入更多歷史。"""
        dt = self._jump_dt_edit.dateTime()
        ts_ms = dt.toMSecsSinceEpoch()

        if not self._loaded_klines:
            return

        oldest_ms = self._loaded_klines[0].open_time
        newest_ms = self._loaded_klines[-1].close_time

        if ts_ms > newest_ms:
            # 未來時間，跳到最新
            self._kline_chart.scroll_to_time(newest_ms)
            return

        if oldest_ms <= ts_ms <= newest_ms:
            self._kline_chart.scroll_to_time(ts_ms)
            self._jump_target_ms = None
        else:
            # 需要載入更早的歷史資料
            self._jump_target_ms = ts_ms
            self._status_lbl.setText("跳轉：載入歷史中…")
            if self._ws_thread:
                self._ws_thread.request_more_history(oldest_ms)

    def _on_realtime_toggled(self, checked: bool) -> None:
        """⚡ 即時按鈕 toggle：開啟後每根收盤 K 棒自動標注。"""
        self._strategy_realtime = checked

    def _on_clear_strategy(self) -> None:
        """清除所有策略標記、訊號與統計。"""
        self._kline_chart.clear_strategy_markers()
        self._strategy_signals = []
        self._strategy_stats_lbl.setVisible(False)
        self._strategy_stats_lbl.setText("")


    def _update_strategy_stats(self) -> None:
        """根據目前 _strategy_signals 計算並顯示統計 label。"""
        if not self._strategy_engine or not self._strategy_signals:
            self._strategy_stats_lbl.setVisible(False)
            return
        stats = self._strategy_engine.compute_stats(self._strategy_signals)
        self._show_strategy_stats_label(stats)

    def _show_strategy_stats_label(self, stats: dict) -> None:
        """將回測統計顯示在工具列 label 上。"""
        n      = stats.get("trades", 0)
        wr     = stats.get("win_rate", 0.0)
        pnl    = stats.get("total_pnl", 0.0)
        opens  = stats.get("open_count", 0)
        pf     = stats.get("profit_factor", 0.0)
        mcl    = stats.get("max_consec_loss", 0)
        mdd    = stats.get("max_drawdown", 0.0)
        ln     = stats.get("long_trades", 0)
        sn     = stats.get("short_trades", 0)

        pnl_sign = "+" if pnl >= 0 else ""
        pf_s = f"{pf:.2f}" if pf != float("inf") else "∞"
        txt = (f"{n} 筆  勝率 {wr:.1f}%  PnL {pnl_sign}{pnl:.2f}%"
               f"  PF {pf_s}  連虧 {mcl}  回撤 {mdd:.2f}%"
               f"  L{ln}/S{sn}")
        if opens:
            txt += f"  (+{opens} 未平倉)"
        self._strategy_stats_lbl.setText(txt)
        self._strategy_stats_lbl.setVisible(True)

    # ══════════════════════════════════════════════════════════════
    # 十字線同步
    # ══════════════════════════════════════════════════════════════

    def _sync_crosshair(self, x_pos: float) -> None:
        """任一右側面板的十字線移動時，同步垂直線至其他面板。"""
        sender = self.sender()
        for chart in self._crosshair_charts:
            if chart is not sender:
                chart.set_crosshair_x(x_pos)
        self._stats_panel.set_crosshair_x(x_pos)

    def _hide_all_crosshairs(self) -> None:
        """滑鼠離開面板時，隱藏所有十字線。"""
        for chart in self._crosshair_charts:
            chart.hide_crosshair()
        self._stats_panel.hide_crosshair()

    # ══════════════════════════════════════════════════════════════
    # 視窗生命週期
    # ══════════════════════════════════════════════════════════════

    def closeEvent(self, event) -> None:  # noqa: N802
        self._heatmap_timer.stop()
        self._flush_timer.stop()
        self._stale_timer.stop()
        if self._history_proc and self._history_proc.isRunning():
            self._history_proc.quit()
            self._history_proc.wait(3000)
        if self._ws_thread:
            self._ws_thread.stop()
            if not self._ws_thread.wait(6000):   # 等最多 6 秒讓執行緒自行結束
                self._ws_thread.terminate()      # 後備：強制終止，避免 QThread destroyed 警告
                self._ws_thread.wait(1000)
        event.accept()
