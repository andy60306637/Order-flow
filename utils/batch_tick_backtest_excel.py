"""
批次執行三年區間 Tick 回測並匯出 Excel。

用法：
  python utils/batch_tick_backtest_excel.py

輸出：
  docs/reports/tick_backtest_y1_20230414_20240413.xlsx
  docs/reports/tick_backtest_y2_20240414_20250413.xlsx
  docs/reports/tick_backtest_y3_20250414_20260413.xlsx
"""
from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backtest.engine import BacktestConfig, simulate_trades
from core import tick_cache, kline_cache
from strategies.wick_reversal_v4 import WickReversalV4Strategy

SYMBOL   = "BTCUSDT"
INTERVAL = "1m"
BAR_MS   = 60_000

# ── 回測參數 ─────────────────────────────────────────────────────────────────
CFG = BacktestConfig(
    initial_capital = 10_000.0,
    leverage        = 20,
    fee_mode        = "自訂",
    custom_fee_rate = 0.00032,   # 0.032%
    slippage_bps    = 0.2,
    compound        = False,     # 固定倉位
    maint_margin    = 0.004,
    max_loss_pct    = 0.02,      # 2% — UI 預設值
)

# ── 三年區間 ─────────────────────────────────────────────────────────────────
INTERVALS = [
    ("y1_20230414_20240413", "BTCUSDT_20230414_20240413"),
    ("y2_20240414_20250413", "BTCUSDT_20240414_20250413"),
    ("y3_20250414_20260413", "BTCUSDT"),
]

OUTPUT_DIR = PROJECT_ROOT / "docs" / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _fmt_ms(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M")


def run_one(label: str, tick_symbol: str) -> dict | None:
    print(f"\n{'='*60}")
    print(f"[{label}] tick_symbol={tick_symbol}")

    # ── 1. 載入 tick 資料 ────────────────────────────────────────────
    info = tick_cache.info(tick_symbol)
    if info is None:
        print(f"  [ERR] tick dataset not found: {tick_symbol}")
        return None
    start_ms = info["start_ms"]
    end_ms   = info["end_ms"]
    print(f"  tick 範圍: {_fmt_ms(start_ms)} ~ {_fmt_ms(end_ms)}")

    ticks = tick_cache.load_range(tick_symbol, start_ms, end_ms)
    if len(ticks) == 0:
        print(f"  [ERR] no tick data")
        return None
    print(f"  tick 數量: {len(ticks):,}")

    # ── 2. 載入對應 K 棒 ─────────────────────────────────────────────
    range_start_ms = (start_ms // BAR_MS) * BAR_MS
    range_end_ms   = (end_ms   // BAR_MS) * BAR_MS
    bt_klines = kline_cache.load_range_as_klines(
        SYMBOL, INTERVAL, range_start_ms, range_end_ms
    )
    if not bt_klines:
        print(f"  [ERR] no kline data")
        return None
    print(f"  K 棒數量: {len(bt_klines):,}")

    # ── 3. 建立 tick_map ─────────────────────────────────────────────
    kline_times = [(k.open_time, k.close_time) for k in bt_klines]
    tick_map = tick_cache.build_bar_map(ticks, kline_times)
    tick_coverage_pct = len(tick_map) / len(bt_klines) * 100
    print(f"  tick 覆蓋率: {tick_coverage_pct:.1f}% ({len(tick_map):,}/{len(bt_klines):,})")

    # ── 4. 執行策略 ──────────────────────────────────────────────────
    strategy = WickReversalV4Strategy()
    strategy.allow_bar_fallback_in_tick_mode = False
    signals = strategy.on_history(bt_klines, tick_map=tick_map)
    print(f"  策略訊號數: {len(signals):,}")

    # ── 5. 模擬回測 ──────────────────────────────────────────────────
    sim_stats = simulate_trades(signals, CFG)
    sim_stats["strategy_name"]     = WickReversalV4Strategy.name
    sim_stats["backtest_start_ms"] = bt_klines[0].open_time  if bt_klines else 0
    sim_stats["backtest_end_ms"]   = bt_klines[-1].open_time if bt_klines else 0
    sim_stats["tick_coverage_pct"] = tick_coverage_pct

    trades = sim_stats.get("trades", 0)
    pnl    = sim_stats.get("total_net_pnl", 0.0)
    wr     = sim_stats.get("win_rate", 0.0)
    print(f"  交易數={trades}, 勝率={wr:.1f}%, 淨利={pnl:.2f} USDT")
    return sim_stats


def export_excel(sim_stats: dict, label: str) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("請先安裝 openpyxl：pip install openpyxl")

    s = sim_stats
    h_font = Font(bold=True, color="FFFFFF")
    h_fill = PatternFill("solid", fgColor="2962FF")
    s_fill = PatternFill("solid", fgColor="1e222d")
    center = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()

    # ── Sheet 1: 摘要 ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "摘要"

    def _pf(v):
        return "∞" if v == float("inf") else round(v, 2)

    def _fmt_ts(ms: int) -> str:
        if not ms:
            return "─"
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d %H:%M")

    param_heads = ["資金(U)", "槓桿", "費率模式", "費率%",
                   "損失上限%", "滑價(bps)", "資金費率/8h",
                   "最終餘額", "報酬率%",
                   "回測起始", "回測結束", "Tick覆蓋%", "無Tick棒數"]
    for col, h in enumerate(param_heads, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = h_font, h_fill, center

    param_vals = [
        s.get("initial_capital", 0),
        s.get("leverage", 0),
        s.get("fee_mode", ""),
        round(s.get("fee_rate", 0) * 100, 4),
        round(s.get("max_loss_pct", 0) * 100, 1),
        s.get("slippage_bps", 0.0),
        s.get("funding_rate", 0.0),
        round(s.get("final_equity", 0.0), 2),
        round(s.get("total_return_pct", 0.0), 2),
        _fmt_ts(s.get("backtest_start_ms", 0)),
        _fmt_ts(s.get("backtest_end_ms", 0)),
        round(s["tick_coverage_pct"], 1) if s.get("tick_coverage_pct") is not None else "─",
        s.get("fallback_bar_count", 0),
    ]
    for col, val in enumerate(param_vals, 1):
        ws1.cell(row=2, column=col, value=val).alignment = center

    sum_heads = ["策略", "交易數", "勝率%", "PF",
                 "淨利(USDT)", "手續費",
                 "最大回撤%", "最大連虧",
                 "平均獲利", "平均虧損",
                 "多單PF", "空單PF",
                 "SL", "TP", "TS", "TD", "TDD"]
    for col, h in enumerate(sum_heads, 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font, c.fill, c.alignment = h_font, s_fill, center

    sum_vals = [
        s.get("strategy_name", ""),
        s.get("trades", 0),
        round(s.get("win_rate", 0.0), 1),
        _pf(s.get("profit_factor", 0.0)),
        round(s.get("total_net_pnl", 0.0), 2),
        round(s.get("total_fees", 0.0), 2),
        round(s.get("max_drawdown_pct", 0.0), 2),
        s.get("max_consec_loss", 0),
        round(s.get("avg_win", 0.0), 2),
        round(s.get("avg_loss", 0.0), 2),
        _pf(s.get("long_profit_factor", 0.0)),
        _pf(s.get("short_profit_factor", 0.0)),
        s.get("sl_count", 0),
        s.get("tp_count", 0),
        s.get("ts_count", 0),
        s.get("td_count", 0),
        s.get("tdd_count", 0),
    ]
    for col, val in enumerate(sum_vals, 1):
        cell = ws1.cell(row=5, column=col, value=val)
        cell.alignment = center
        if col == 5:
            net = s.get("total_net_pnl", 0.0)
            cell.font = Font(color="26A69A" if net >= 0 else "EF5350")

    for col in range(1, max(len(param_heads), len(sum_heads)) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # ── Sheet 2: 交易明細 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("交易明細")
    trade_heads = ["#", "方向", "入場時間", "入場價", "出場類型",
                   "出場價", "數量", "手續費",
                   "資金費", "淨利(USDT)", "餘額"]
    trade_heads += [
        "side", "wick_type", "session_hour", "atr_percentile",
        "trend_regime", "entry_delay_bars", "k0_range_atr",
        "wick_volume_ratio", "zoom_delta_eff", "MAE", "MFE",
        "mae_r", "mfe_r", "entry_risk", "trailing_stop_mode",
        "final_trade_delta",
    ]
    for col, h in enumerate(trade_heads, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = h_font, h_fill, center

    trade_list = s.get("trade_list", [])
    active = [t for t in trade_list if not t.get("skipped")]
    for i, t in enumerate(active, 1):
        dir_txt  = "做多" if t["dir"] == "long" else "做空"
        pv       = t.get("net_pnl", 0.0)
        ets      = t.get("entry_time", 0)
        time_str = _fmt_ts(ets) if ets else "─"
        row_vals = [
            i, dir_txt, time_str,
            round(t.get("entry", 0), 4),
            t.get("exit_label", ""),
            round(t.get("exit", 0), 4),
            round(t.get("qty", 0), 6),
            round(t.get("total_fee", 0), 2),
            round(t.get("funding_cost", 0), 2),
            round(pv, 2),
            round(t.get("equity_after", 0), 2),
        ]
        row_vals += [
            t.get("side", ""),
            t.get("wick_type", ""),
            t.get("session_hour", ""),
            t.get("atr_percentile", ""),
            t.get("trend_regime", ""),
            t.get("entry_delay_bars", ""),
            t.get("k0_range_atr", ""),
            t.get("wick_volume_ratio", ""),
            t.get("zoom_delta_eff", ""),
            t.get("MAE", ""),
            t.get("MFE", ""),
            t.get("mae_r", ""),
            t.get("mfe_r", ""),
            t.get("entry_risk", ""),
            t.get("trailing_stop_mode", ""),
            t.get("final_trade_delta", ""),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=i + 1, column=col, value=val)
            cell.alignment = center
            if col == 10:
                cell.font = Font(color="26A69A" if pv >= 0 else "EF5350")

    for col in range(1, len(trade_heads) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # ── Sheet 3: 分組統計 ──────────────────────────────────────────────
    def _write_group_sheet(ws, title: str, data: dict) -> None:
        grp_heads = ["分組", "交易數", "勝率%", "PF", "淨利(U)", "毛利(U)",
                     "手續費(U)", "平均損益", "平均R", "平均MAE_R", "平均MFE_R"]
        for col, h in enumerate(grp_heads, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = h_font, h_fill, center
        for row, (key, gs) in enumerate(data.items(), 2):
            vals = [
                str(key),
                gs.get("trades", 0),
                round(gs.get("win_rate", 0.0), 1),
                _pf(gs.get("pf", 0.0)),
                round(gs.get("net_pnl", 0.0), 2),
                round(gs.get("gross_pnl", 0.0), 2),
                round(gs.get("fees", 0.0), 2),
                round(gs.get("avg_pnl", 0.0), 2),
                round(gs.get("avg_R", 0.0), 3),
                round(gs.get("avg_mae_r", 0.0), 3),
                round(gs.get("avg_mfe_r", 0.0), 3),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = center
                if col == 5:
                    net = gs.get("net_pnl", 0.0)
                    c.font = Font(color="26A69A" if net >= 0 else "EF5350")
        for col in range(1, len(grp_heads) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    ws3 = wb.create_sheet("Exit統計")
    _write_group_sheet(ws3, "Exit統計", s.get("exit_stats", {}))

    ws4 = wb.create_sheet("多空統計")
    _write_group_sheet(ws4, "多空統計", s.get("side_stats", {}))

    ws5 = wb.create_sheet("Side+Regime")
    _write_group_sheet(ws5, "Side+Regime", s.get("regime_side_stats", {}))

    ws6 = wb.create_sheet("WickType統計")
    _write_group_sheet(ws6, "WickType統計", {
        **s.get("wick_type_stats", {}), **s.get("side_wick_type_stats", {})
    })

    ws7 = wb.create_sheet("Session Hour")
    _write_group_sheet(ws7, "Session Hour", s.get("session_hour_stats", {}))

    ws8 = wb.create_sheet("EntryDelay")
    _write_group_sheet(ws8, "EntryDelay", s.get("entry_delay_bars_stats", {}))

    out_path = OUTPUT_DIR / f"tick_backtest_{label}.xlsx"
    wb.save(out_path)
    print(f"  [OK] saved: {out_path}")
    return out_path


def main():
    results = []
    for label, tick_symbol in INTERVALS:
        sim_stats = run_one(label, tick_symbol)
        if sim_stats is None:
            print(f"  跳過 {label}")
            continue
        out = export_excel(sim_stats, label)
        results.append((label, out))

    print(f"\n{'='*60}")
    print(f"完成！共匯出 {len(results)} 個 Excel 檔案：")
    for label, path in results:
        print(f"  [{label}] {path}")


if __name__ == "__main__":
    main()
